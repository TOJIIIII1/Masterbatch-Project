import pandas as pd
import psycopg2
import ttkbootstrap as ttk
import tkinter as tk
from tkinter import messagebox
from tkinter.filedialog import asksaveasfilename
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font


class Notes:
    def __init__(self):
        self.conn = None
        self.cursor = None
        self.product_kinds = []  # List to hold product kinds
        self.connect_db()

    def connect_db(self):
        """Establish a connection to the PostgreSQL database and create a cursor."""
        try:
            if self.conn is None or self.conn.closed != 0:
                self.conn = psycopg2.connect(
                    host="localhost",
                    port=5431,
                    dbname="Inventory",
                    user="postgres",
                    password="newpassword"
                )
                self.cursor = self.conn.cursor()
                print("Database connection established.")
                self.fetch_product_kinds()  # Fetch product kinds when the connection is made
        except Exception as e:
            print(f"Error connecting to database: {e}")

    def fetch_product_kinds(self):
        """Fetch the available product kinds from the database."""
        try:
            query = "SELECT DISTINCT product_kind FROM notes"
            self.cursor.execute(query)
            self.product_kinds = [row[0] for row in self.cursor.fetchall()]
            if not self.product_kinds:
                self.product_kinds = ["Default Product Kind 1", "Default Product Kind 2"]
        except Exception as e:
            print(f"Error fetching product kinds: {e}")
            self.product_kinds = ["Default Product Kind 1", "Default Product Kind 2"]

    def fetch_data_from_notes(self):
        """Fetch the latest data from the Notes table, excluding deleted rows."""
        try:
            query = """
                SELECT product_code, 
                       lot_number, 
                       product_kind 
                FROM notes
                WHERE deleted = FALSE;
            """
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except Exception as e:
            print(f"Error fetching data: {e}")
            return []

    def update_treeview(self, table, data, columns):
        """Update the Treeview with new data."""
        for row in table.get_children():
            table.delete(row)  # Remove all existing rows

        for row_data in data:
            table.insert('', 'end', values=row_data)  # Insert the new rows

    def display(self, parent_frame):
        """Display the Notes page."""
        for widget in parent_frame.winfo_children():
            widget.destroy()

        header = ttk.Label(
            parent_frame, text="Notes Management", style="Custom.TLabel", font=("Arial", 20, "bold")
        )
        header.grid(row=0, column=0, columnspan=3, pady=10, sticky="ew")

        table_frame = ttk.Frame(parent_frame)
        table_frame.grid(row=1, column=0, columnspan=3, padx=10, pady=10, sticky="nsew")

        column_names = ["Product Code", "Lot Number", "Product Kind"]
        notes_tree = ttk.Treeview(
            table_frame, columns=column_names, show="headings", height=15
        )
        for col_name in column_names:
            notes_tree.heading(col_name, text=col_name)
            notes_tree.column(col_name, width=150, anchor="center")
        notes_tree.pack(fill="both", expand=True)

        data_notes = self.fetch_data_from_notes()
        self.update_treeview(notes_tree, data_notes, column_names)

        entry_frame = ttk.Frame(parent_frame)
        entry_frame.grid(row=2, column=0, columnspan=3, padx=10, pady=10)

        labels = ["Product Code", "Lot Number", "Product Kind"]
        self.notes_entries = []

        for i, label_text in enumerate(labels):
            ttk.Label(entry_frame, text=label_text, font=("Arial", 12)).grid(
                row=i, column=0, padx=10, pady=5, sticky="e"
            )
            if label_text == "Product Kind":
                self.product_kinds = ["MB", "DC"]  # Add options here
                combobox = ttk.Combobox(entry_frame, values=self.product_kinds, width=30, state="readonly")
                combobox.grid(row=i, column=1, padx=10, pady=5, sticky="w")
                self.notes_entries.append(combobox)

                # Set default selection
                combobox.set(self.product_kinds[0])  # Default to 'MB'
            else:
                entry = ttk.Entry(entry_frame, width=30)
                entry.grid(row=i, column=1, padx=10, pady=5, sticky="w")
                self.notes_entries.append(entry)

        button_frame = ttk.Frame(parent_frame)
        button_frame.grid(row=3, column=0, columnspan=3, pady=10)

        ttk.Button(
            button_frame, text="Add", command=lambda: self.add_row_notes(notes_tree)
        ).grid(row=0, column=0, padx=10)
        ttk.Button(
            button_frame, text="Update", command=lambda: self.update_row_notes(notes_tree)
        ).grid(row=0, column=1, padx=10)
        ttk.Button(
            button_frame, text="Delete", command=lambda: self.delete_row_notes(notes_tree)
        ).grid(row=0, column=2, padx=10)
        ttk.Button(
            button_frame, text="Clear", command=lambda: self.clear_entries(notes_tree, clear_ui_only=True),
        ).grid(row=0, column=3, padx=10)

        # Export Button
        ttk.Button(
            button_frame, text="Export to Excel", command=lambda: self.export_to_excel(notes_tree)
        ).grid(row=0, column=4, padx=10)

    def add_row_notes(self, table):
        """Add a new row to the Notes table."""
        product_code = self.notes_entries[0].get().strip().upper()
        lot_number = self.notes_entries[1].get().strip().upper()
        product_kind = self.notes_entries[2].get().strip().upper()

        if not (product_code and lot_number and product_kind):
            messagebox.showerror("Error", "All fields must be filled out.")
            return

        try:
            query = """
                INSERT INTO notes (product_code, lot_number, product_kind) 
                VALUES (%s, %s, %s)
            """
            self.cursor.execute(query, (product_code, lot_number, product_kind))
            self.conn.commit()
            self.update_treeview(table, self.fetch_data_from_notes(), ["Product Code", "Lot Number", "Product Kind"])
        except Exception as e:
            messagebox.showerror("Error", f"Error adding note: {e}")
            print(f"{e}")

    def update_row_notes(self, table):
        """Update an existing row in the Notes table."""
        try:
            self.connect_db()  # Ensure the connection is open

            # Get selected row
            selected_item = table.selection()
            if not selected_item:
                messagebox.showwarning("No Selection", "Please select a row to update.")
                return

            # Retrieve values from entry fields
            product_code = self.notes_entries[0].get()
            lot_number = self.notes_entries[1].get()
            product_kind = self.notes_entries[2].get()

            if not (product_code and lot_number and product_kind):
                messagebox.showerror("Error", "All fields must be filled out.")
                return

            # Ensure product_code is treated as a string (even if it's a number)
            product_code = str(product_code)  # Force product_code to be a string

            # Get the current product_code of the selected row (used for the WHERE clause)
            current_product_code = table.item(selected_item)['values'][0]

            # Ensure current_product_code is also treated as a string
            current_product_code = str(current_product_code)

            # Construct the update query
            query = """
                UPDATE notes 
                SET product_code = %s, lot_number = %s, product_kind = %s
                WHERE product_code = %s
            """

            # Execute the update query
            self.cursor.execute(query, (product_code, lot_number, product_kind, current_product_code))
            self.conn.commit()

            # Show success message
            messagebox.showinfo("Success", f"Row with Product Code: {current_product_code} updated successfully.")

            # Refresh the Treeview to show the updated data
            data_from_notes = self.fetch_data_from_notes()
            self.update_treeview(table, data_from_notes, ["Product Code", "Lot Number", "Product Kind"])

        except Exception as e:
            messagebox.showerror("Error", f"Error while updating row in Table Notes: {e}")
            self.conn.rollback()  # Rollback the transaction if there's an error

    def delete_row_notes(self, table):
        try:
            self.connect_db()  # Ensure the connection is open

            # Get selected row
            selected_item = table.selection()
            if not selected_item:
                messagebox.showwarning("No Selection", "Please select a row to delete.")
                return

            # Get the reference_no (used as the unique identifier) of the selected row
            product_code = table.item(selected_item)['values'][0]  # Assuming first column is reference_no (text)

            # Debugging: Check the reference_no value
            print(f"Selected Product Code: {product_code}")

            # Ask for confirmation
            confirm = messagebox.askyesno("Confirm Delete",
                                          f"Are you sure you want to delete the row with Reference No: {product_code}?")
            if not confirm:
                return

            # Ensure reference_no is treated as text (explicitly casting to TEXT in SQL)
            query = "DELETE FROM notes WHERE product_code = %s::TEXT"

            # Debugging: Check the query and parameter
            print(f"SQL Query: {query} - Product Code: {product_code}")

            # Execute the delete query
            self.cursor.execute(query, (product_code,))  # Treat reference_no as text
            self.conn.commit()

            messagebox.showinfo("Success", f"Row with Reference No: {product_code} deleted successfully.")

            # Refresh the Treeview to show the updated data
            data_from_notes = self.fetch_data_from_notes()
            self.update_treeview(table, data_from_notes,
                                 ["Product Code", "Lot Number", "Product Kind"])

        except Exception as e:
            messagebox.showerror("Error", f"Error while deleting row in Table 1: {e}")
            self.conn.rollback()  # Rollback the transaction if there's an error

    def clear_entries(self, table, clear_ui_only=True):
        """
        Mark all rows as deleted (flagging rows in the UI and database).
        The `clear_ui_only` flag determines if only the UI should be cleared or if other actions are required.
        """
        try:
            # Confirm if the user wants to clear all rows from the UI
            print("clear_entries method called")
            confirm = messagebox.askyesno("Confirm Clear", "Are you sure you want to mark all rows as deleted?")
            if not confirm:
                return

            if clear_ui_only:
                # Clear all rows from the Treeview (UI)
                table.delete(*table.get_children())  # This removes all the rows in the table.
                messagebox.showinfo("Success", "All rows have been cleared from the table (UI only).")

                # Mark rows as deleted in the database
                query = "UPDATE notes SET deleted = TRUE WHERE deleted = FALSE;"
                self.cursor.execute(query)
                self.conn.commit()
                messagebox.showinfo("Database Update", "All rows have been marked as deleted in the database.")

            # Reload the table to reflect the changes after the update
            self.update_treeview(table, self.fetch_data_from_notes(), ["Product Code", "Lot Number", "Product Kind"])

        except Exception as e:
            messagebox.showerror("Error", f"Error while clearing rows: {e}")
            self.conn.rollback()  # Rollback the transaction if there's an error

    def export_to_excel(self, treeview):
        """Export the data from the Treeview to an Excel file with custom rows and additional database sheets."""

        # Calculate yesterday's date
        yesterday_date = (datetime.now() - timedelta(days=1)).strftime("%m/%d/%Y")

        data = []
        for row in treeview.get_children():
            data.append(treeview.item(row)["values"])

        if not data:
            messagebox.showerror("Error", "No data available to export.")
            return

        file_path = asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return

        try:
            # Create custom headers
            top_row_1 = ["Daily Ending Inventory Report From:", f"{yesterday_date}", "", ""]
            top_row_2 = ["List of Batches Included in Report", "", "", ""]
            top_row_3 = ["MASTERBATCH", "", "", ""]
            top_row_4 = ["PRODUCT CODE", "LOT#", "Product Kind", ""]

            # Insert headers at the beginning of the data list
            data.insert(0, top_row_1)
            data.insert(1, top_row_2)
            data.insert(2, top_row_3)
            data.insert(3, top_row_4)

            # Define column names for Treeview data
            columns = ["Product Code", "Lot Number", "Product Kind"]

            # Convert the remaining data (excluding first 4 rows) to a DataFrame
            df = pd.DataFrame(data[4:], columns=columns)

            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Write the data (starting from row 5 in Excel)
                df.to_excel(writer, index=False, header=False, startrow=4, sheet_name="NOTES")
                sheet = writer.sheets['NOTES']

                # Write header rows manually
                for row_num, row_data in enumerate([top_row_1, top_row_2, top_row_3, top_row_4], start=1):
                    for col_num, value in enumerate(row_data, start=1):
                        sheet.cell(row=row_num, column=col_num, value=value)

                # Function to auto-adjust column widths
                def auto_adjust_column_width(sheet):
                    for col in sheet.columns:
                        max_length = 0
                        col_letter = get_column_letter(col[0].column)  # Get column letter (A, B, C, etc.)
                        for cell in col:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        sheet.column_dimensions[col_letter].width = max_length + 2  # Add some padding

                # Export Warehouse 1 data
                query_whse1 = """
                    SELECT DISTINCT
                        m.material_code,   
                        'NaN'::FLOAT AS number_of_bags,  -- PostgreSQL only (for FLOAT/NUMERIC)
                        'NaN'::FLOAT AS qty_per_packing,
                        'NaN'::FLOAT AS whse1_excess,
                        ROUND(COALESCE(mt.total_quantity, 0)::NUMERIC, 2) AS totals,  -- Cast to NUMERIC and round to 2 decimal places
                        CASE
                            WHEN COALESCE(mt.status, 'Good') = 'Good' THEN NULL  -- Make status blank if it's 'Good'
                            ELSE COALESCE(mt.status, 'Good')  -- Otherwise, keep the original status
                        END AS status  
                    FROM 
                        material_codes AS m
                    LEFT JOIN 
                        wh1_material_code_totals AS mt ON m.mid = mt.ID  
                    LEFT JOIN 
                        wh1_transfer_form AS wtf ON m.mid = wtf.material_code
                    LEFT JOIN 
                        wh1_receiving_report AS wrr ON m.mid = wrr.material_code
                    WHERE 
                        COALESCE(mt.total_quantity, 0) > 0  
                    ORDER BY 
                        m.material_code, status ASC;
                """

                self.cursor.execute(query_whse1)
                data_whse1 = self.cursor.fetchall()
                df_whse1 = pd.DataFrame(data_whse1,
                                        columns=[f"{yesterday_date}", "No of Bags", "Qty per Packing",
                                                 "WHSE #1 - Excess", "Total", "Status"])
                df_whse1.to_excel(writer, sheet_name="WHSE1", index=False)
                auto_adjust_column_width(writer.sheets["WHSE1"])  # Auto-adjust column widths

                # Export Warehouse 2 data
                query_whse2 = """
                    SELECT DISTINCT
                        m.material_code,   
                        'NaN'::FLOAT AS number_of_bags,  -- PostgreSQL only (for FLOAT/NUMERIC)
                        'NaN'::FLOAT AS qty_per_packing,
                        'NaN'::FLOAT AS whse1_excess,
                        ROUND(COALESCE(mt.total_quantity, 0)::NUMERIC, 2) AS totals,  -- Cast to NUMERIC and round to 2 decimal places
                        CASE
                            WHEN COALESCE(mt.status, 'Good') = 'Good' THEN NULL  -- Make status blank if it's 'Good'
                            ELSE COALESCE(mt.status, 'Good')  -- Otherwise, keep the original status
                        END AS status  
                    FROM 
                        wh2_material_codes AS m
                    LEFT JOIN 
                        wh2_material_code_totals AS mt ON m.mid = mt.ID  
                    LEFT JOIN 
                        wh2_transfer_form AS wtf ON m.mid = wtf.material_code
                    LEFT JOIN 
                        wh2_receiving_report AS wrr ON m.mid = wrr.material_code
                    WHERE 
                        COALESCE(mt.total_quantity, 0) > 0  
                    ORDER BY 
                        m.material_code, status ASC;
                """

                self.cursor.execute(query_whse2)
                data_whse2 = self.cursor.fetchall()
                df_whse2 = pd.DataFrame(data_whse2,
                                        columns=[f"{yesterday_date}", "No of Bags", "Qty per Packing",
                                                 "WHSE #1 - Excess", "Total", "Status"])
                df_whse2.to_excel(writer, sheet_name="WHSE2", index=False)
                auto_adjust_column_width(writer.sheets["WHSE2"])  # Auto-adjust column widths

                # Export Warehouse 4 data
                query_whse4 = """
                    SELECT DISTINCT
                        m.material_code,   
                        'NaN'::FLOAT AS number_of_bags,  -- PostgreSQL only (for FLOAT/NUMERIC)
                        'NaN'::FLOAT AS qty_per_packing,
                        'NaN'::FLOAT AS whse1_excess,
                        ROUND(COALESCE(mt.total_quantity, 0)::NUMERIC, 2) AS totals,  -- Cast to NUMERIC and round to 2 decimal places
                        CASE
                            WHEN COALESCE(mt.status, 'Good') = 'Good' THEN NULL  -- Make status blank if it's 'Good'
                            ELSE COALESCE(mt.status, 'Good')  -- Otherwise, keep the original status
                        END AS status  
                    FROM 
                        wh4_material_codes AS m
                    LEFT JOIN 
                        wh4_material_code_totals AS mt ON m.mid = mt.ID  
                    LEFT JOIN 
                        wh4_transfer_form AS wtf ON m.mid = wtf.material_code
                    LEFT JOIN 
                        wh4_receiving_report AS wrr ON m.mid = wrr.material_code
                    WHERE 
                        COALESCE(mt.total_quantity, 0) > 0  
                    ORDER BY 
                        m.material_code, status ASC;
                """
                self.cursor.execute(query_whse4)
                data_whse4 = self.cursor.fetchall()
                df_whse4 = pd.DataFrame(data_whse4,
                                        columns=[f"{yesterday_date}", "No of Bags", "Qty per Packing",
                                                 "WHSE #4 - Excess", "Total", "Status"])
                df_whse4.to_excel(writer, sheet_name="WHSE4", index=False)
                auto_adjust_column_width(writer.sheets["WHSE4"])  # Auto-adjust column widths

            messagebox.showinfo("Success", f"Data successfully exported to {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Error exporting data: {e}")
            print(f"{e}")





